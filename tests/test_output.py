"""Tests for JSON and Excel serializers."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from baltic_scraper.output import (
    _safe_sheet_name,
    write_excel,
    write_json,
)

SAMPLE = {
    "VLCC (Dirty Tanker)": {
        "TD02: Ras Tanura to Singapore": {
            "Time Charter Equivalent (TCE) Outcome": {
                "Time Charter Equivalent ($/day)": {
                    "your_outcome": "$46,997.73",
                    "baltic_outcome": "$46,997.73",
                    "difference": "$0.00",
                }
            },
            "Income": {
                "Total Voyage Days": {
                    "your_outcome": "34.706",
                    "baltic_outcome": "34.706",
                    "difference": "0.000",
                }
            },
        }
    },
    "Suezmax (Dirty Tanker)": {
        "TD06: CPC Marine Terminal to Augusta": {
            "Ports": {
                "Load Port Charges": {
                    "your_outcome": "$1.00",
                    "baltic_outcome": "$1.00",
                    "difference": "$0.00",
                }
            }
        }
    },
}


def test_write_json_roundtrip(tmp_path: Path) -> None:
    """JSON output is valid and matches the input data."""
    out = tmp_path / "out.json"
    write_json(SAMPLE, out)
    loaded = json.loads(out.read_text(encoding="utf-8"))
    assert loaded == SAMPLE


def test_write_excel_one_sheet_per_route(tmp_path: Path) -> None:
    """Excel has one worksheet per route code."""
    out = tmp_path / "out.xlsx"
    write_excel(SAMPLE, out)
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"TD02", "TD06"}


def test_write_excel_contains_values(tmp_path: Path) -> None:
    """Scraped values appear in the worksheet cells."""
    out = tmp_path / "out.xlsx"
    write_excel(SAMPLE, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["TD02"]
    all_values = {c.value for row in ws.iter_rows() for c in row}
    assert "$46,997.73" in all_values
    assert "Time Charter Equivalent ($/day)" in all_values
    assert "VLCC (Dirty Tanker)" in all_values


def test_write_excel_empty_data(tmp_path: Path) -> None:
    """Empty data still yields a valid workbook with a placeholder sheet."""
    out = tmp_path / "empty.xlsx"
    write_excel({}, out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["No data"]


def test_write_excel_skips_non_dict(tmp_path: Path) -> None:
    """Non-dict route values (e.g. list-mode) are ignored gracefully."""
    out = tmp_path / "list.xlsx"
    write_excel({"VLCC": ["TD02", "TD03"]}, out)  # list-mode shape
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["No data"]


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("TD02", "TD02"),
        ("a/b:c*d?e[f]g", "abcdefg"),
        ("x" * 40, "x" * 31),
    ],
)
def test_safe_sheet_name(raw: str, expected: str) -> None:
    """Illegal characters are stripped and length is capped at 31."""
    assert _safe_sheet_name(raw) == expected
