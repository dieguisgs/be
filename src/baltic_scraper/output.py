"""
Output serializers: JSON and Excel (.xlsx) for scraped TCE data.

Excel layout per sheet (one sheet per vessel-class + route):

    Row 1  : Vessel class name  (dark-blue header, merged)
    Row 2  : Route name         (dark-blue header, merged)
    Row 3  : blank
    Row N  : Section header     (dark-blue, merged A-D)
    Row N+1: Column labels      (Name | Your Outcome | Baltic Outcome | Difference)
    Row N+2+: Data rows
    ...    : blank row between sections
"""

from __future__ import annotations

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False

# Colour constants (ARGB hex, no leading #)
_DARK_BLUE = "FF0D3557"
_TEAL = "FF00838F"
_WHITE = "FFFFFFFF"
_LIGHT_GREY = "FFF5F5F5"

SectionData = dict[str, dict[str, str]]
RouteData = dict[str, SectionData]
ScraperResult = dict[str, RouteData]


# ── JSON output ────────────────────────────────────────────────────────────────

def write_json(data: ScraperResult, path: Path) -> None:
    """
    Serialise *data* to a pretty-printed JSON file.

    Parameters
    ----------
    data : ScraperResult
        Nested dict returned by the orchestrator:
        ``{vessel_class: {route: {section: {name: {your_outcome, …}}}}}``.
    path : Path
        Destination file path.  Parent directories must exist.
    """
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _fill(hex_argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_argb)


def _font(bold: bool = False, colour: str = _WHITE, size: int = 11) -> Font:
    return Font(bold=bold, color=colour, size=size)


def _safe_sheet_name(text: str, max_len: int = 31) -> str:
    """
    Return a valid Excel sheet name (max 31 chars, no illegal characters).

    Parameters
    ----------
    text : str
        Raw sheet name candidate.
    max_len : int
        Maximum allowed length (Excel limit is 31).

    Returns
    -------
    str
        Sanitised sheet name.
    """
    illegal = r"\/:*?[]"
    for ch in illegal:
        text = text.replace(ch, "")
    return text[:max_len]


# ── Public Excel writer ────────────────────────────────────────────────────────

def write_excel(data: ScraperResult, path: Path) -> None:  # noqa: PLR0912, PLR0915
    """
    Write *data* to an Excel workbook with one sheet per **route**.

    If the same route appears across multiple vessel classes, all vessel-class
    blocks are written sequentially within that single sheet, each prefixed by
    a vessel-class header row.

    Parameters
    ----------
    data : ScraperResult
        Nested dict: ``{vessel_class: {route: {section: {name: row_dict}}}}``.
    path : Path
        Destination ``.xlsx`` file path.

    Raises
    ------
    ImportError
        If ``openpyxl`` is not installed.
    """
    if not _OPENPYXL_AVAILABLE:
        msg = (
            "openpyxl is required for Excel output. "
            "Install it with: pip install openpyxl"
        )
        raise ImportError(msg)

    # Re-index: route -> [(vessel_class, sections), ...]
    by_route: dict[str, list[tuple[str, RouteData]]] = {}
    for vessel_class, routes in data.items():
        if not isinstance(routes, dict):
            continue
        for route, sections in routes.items():
            if not isinstance(sections, dict):
                continue
            by_route.setdefault(route, []).append((vessel_class, sections))

    wb = Workbook()
    wb.remove(wb.active)

    seen_names: set[str] = set()
    col_widths = [40, 22, 22, 18]
    n_cols = len(col_widths)
    col_labels = ["Name", "Your Outcome", "Baltic Outcome", "Difference"]

    def _merge_header(ws, row: int, text: str, fill_hex: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.fill = _fill(fill_hex)
        cell.font = _font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    for route, vc_list in by_route.items():
        route_code = route.split(":")[0].strip() if ":" in route else route[:28]
        sheet_name = _safe_sheet_name(route_code)
        # Handle duplicates (unlikely but safe)
        original = sheet_name
        counter = 2
        while sheet_name in seen_names:
            sheet_name = _safe_sheet_name(f"{original}_{counter}")
            counter += 1
        seen_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        current_row = 1

        # Route name as top header (spans all vessel-class blocks)
        _merge_header(ws, current_row, route, _DARK_BLUE)
        current_row += 2

        for vessel_class, sections in vc_list:
            # Vessel-class subheader
            _merge_header(ws, current_row, vessel_class, _DARK_BLUE)
            current_row += 1

            for section_name, rows in sections.items():
                # Section header (slightly lighter shade — reuse teal)
                _merge_header(ws, current_row, section_name, _TEAL)
                current_row += 1

                for col_idx, label in enumerate(col_labels, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=label)
                    cell.fill = _fill(_DARK_BLUE)
                    cell.font = _font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                current_row += 1

                for i, (row_name, values) in enumerate(rows.items()):
                    bg = _LIGHT_GREY if i % 2 == 0 else _WHITE
                    for col_idx, val in enumerate(
                        [
                            row_name,
                            values.get("your_outcome", ""),
                            values.get("baltic_outcome", ""),
                            values.get("difference", ""),
                        ],
                        start=1,
                    ):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.fill = _fill(bg)
                        cell.font = Font(color="FF000000")
                    current_row += 1

                current_row += 1  # blank between sections

            current_row += 1  # blank between vessel classes

        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    if not wb.sheetnames:
        wb.create_sheet("No data")

    wb.save(path)
